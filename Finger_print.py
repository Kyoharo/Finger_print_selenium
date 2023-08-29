#Finger_print
from selenium  import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import openpyxl
import datetime
from send_mail import send_email_with_attendance,email_to_helpdesk,email_to_soc,email_to_noc
import concurrent.futures

now = datetime.datetime.now()
current_data = f"{str(now.year)}/{str(now.month).zfill(2)}/{str(now.day).zfill(2)}"
# Format the date
class InstaBot:
    def __init__(self,username,password):  
        #webdriver
        try:
            # serv_obj = service= Service(r'\\10.199.199.35\soc team\Abdelrahman Ataa\Finger_print\geckodriver.exe')
            serv_obj = service= Service(r'geckodriver')
            ops=webdriver.FirefoxOptions()
            self.driver = webdriver.Firefox(service=serv_obj,options=ops)
            self.driver.get("http://10.70.201.21/cgi-bin/login.cgi?command=0")
            self.driver.implicitly_wait(5)
        except Exception as e :
            pass
            # messagebox.showerror("Error", f"There is problem with geckodriver.exe or problem with internet \n  geckodriver هنالك مشكلة بالانترنت او بملف ")
            #try to use chromedriver instead
        if not hasattr(self, 'driver'):
            try:
                serv_obj = Service(r'chromedriver')
                ops = webdriver.ChromeOptions()
                self.driver = webdriver.Chrome(service=serv_obj, options=ops)
                self.driver.get("http://10.70.201.21/cgi-bin/login.cgi?command=0")
                self.driver.implicitly_wait(5)
            except Exception as e:
                # Handle exception
                pass
                return

        try:
            self.username= username
            self.password= password
            #username
            self.driver.find_element(By.XPATH, "//input[@name='loginName']").send_keys(username)
            #password
            self.driver.find_element(By.XPATH, "//input[@name='loginPass']").send_keys(password)
            #login
            self.driver.find_element(By.XPATH, "//input[@value='Login']").click()
            sleep(2)
        except Exception as e :
            print("exception_page1")
            try:
                self.driver.execute_script(f"window.open('http://10.70.201.21/cgi-bin/login.cgi?command=0', '_blank');")
                self.driver.switch_to.window(self.driver.window_handles[-1])  # Switch to the newly opened tab
                sleep(2)
                self.driver.find_element(By.XPATH, "//input[@name='loginName']").send_keys(username)
                #password
                self.driver.find_element(By.XPATH, "//input[@name='loginPass']").send_keys(password)
                #login
                self.driver.find_element(By.XPATH, "//input[@value='Login']").click()
                sleep(5)
            except Exception:
                self.driver.quit()
            return

    def get_attendance(self):
        try:
            self.driver.switch_to.frame("menu") 
            sleep(2)
            self.driver.find_element(By.XPATH, "//a[normalize-space()='View Attendance Report']").click()
            #-----------------------
            sleep(1)
            self.driver.switch_to.default_content()
            self.driver.switch_to.frame("content") 
            sleep(2)
            data = self.driver.find_element(By.XPATH, "//tbody/tr[2]/td[3]").text
            time = self.driver.find_element(By.XPATH, "//tbody/tr[2]/td[4]").text
            trigger = self.driver.find_element(By.XPATH, "//tbody/tr[2]/td[5]").text
            data = {"Date":data,
                "Time":time,
                "Trigger":trigger}
            self.driver.quit()
            return data
        except Exception:
            try:
                self.driver.execute_script(f"window.open('http://10.70.201.21/admin.html', '_blank');")
                self.driver.switch_to.window(self.driver.window_handles[-1])  # Switch to the newly opened tab
                sleep(2)
                self.driver.switch_to.frame("menu") 
                sleep(2)
                self.driver.find_element(By.XPATH, "//a[normalize-space()='View Attendance Report']").click()
                #-----------------------
                sleep(1)
                self.driver.switch_to.default_content()
                self.driver.switch_to.frame("content") 
                sleep(2)
                data = self.driver.find_element(By.XPATH, "//tbody/tr[2]/td[3]").text
                time = self.driver.find_element(By.XPATH, "//tbody/tr[2]/td[4]").text
                trigger = self.driver.find_element(By.XPATH, "//tbody/tr[2]/td[5]").text
                data = {"Date":data,
                    "Time":time,
                    "Trigger":trigger}
                self.driver.quit()
                return data
            except Exception:
                self.driver.quit()
            finally:
                self.driver.quit()

                

        




# sheet_path  = r'\\10.199.199.35\soc team\Abdelrahman Ataa\Finger_print\finger_print.xlsx'
sheet_path  = r'finger_print.xlsx'

wb = openpyxl.load_workbook(sheet_path)
sheet_names = wb.sheetnames



def get_data(id,passwd,attendance,username):
    mybot = InstaBot(id, passwd)
    mydata = mybot.get_attendance()
    if mydata != None:
        print(mydata)
        if mydata['Date'] == current_data:
            attendance.append({
            'Username': username,
            'Trigger': mydata['Trigger'],
            'Time': mydata['Time']
        })
        return 1

def soc_fun(sheet_name):
    attendance = []
    sheet = wb[sheet_name]
    max_rows = sheet.max_row + 1
    for row in range(2, max_rows):
        username = str(sheet.cell(row, 1).value)
        id = str(sheet.cell(row, 2).value)
        passwd = str(sheet.cell(row, 3).value)
        data = get_data(id,passwd,attendance,username)
        while data == None:
            data = get_data(id,passwd,attendance,username)
    return attendance

if __name__ == "__main__":
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        future_soc = executor.submit(soc_fun, "SOC")
        future_noc = executor.submit(soc_fun, "NOC")

        # Wait for both tasks to complete
        concurrent.futures.wait([future_soc, future_noc])

        # Retrieve the results
        result_soc = future_soc.result()
        result_noc = future_noc.result()

    print("Processing of both sheets 'SOC' and 'NOC' has completed.")
    print("Results for 'SOC':", result_soc)
    print("Results for 'NOC':", result_noc)
IN_count_soc = 0
OUT_count_soc = 0
IN_count_noc = 0
OUT_count_noc = 0
for item in result_soc:
    if item['Trigger'] == 'IN':
        IN_count_soc += 1
    else:
        OUT_count_soc += 1
for item in result_noc:
    if item['Trigger'] == 'IN':
        IN_count_noc += 1
    else:
        OUT_count_noc += 1

send_email_with_attendance(result_soc,email_to_soc,"SOC Team",IN_count_soc,OUT_count_soc)
send_email_with_attendance(result_noc,email_to_noc,"NOC Team",IN_count_noc,OUT_count_noc)